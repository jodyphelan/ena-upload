"""
A library for uploading fastq data to the 
European Nucleotide Archive (ENA).
"""

__version__ = "0.1.0"

#! /usr/bin/env python
import sys
import csv
import argparse
import os
import toml
from ftplib import FTP
from tqdm import tqdm
import fastq_files as ff
from typing import List, Union
from fastq_files import SingleSample, PairedSample
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl import Workbook, worksheet
from openpyxl.utils import quote_sheetname

__package_dir__ = os.path.dirname(os.path.abspath(__file__))
alphabet = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"

def cli():
    # create the top-level parser
    parser = argparse.ArgumentParser(description='ENA Uploader')
    parser.add_argument('--version', action='version', version='%(prog)s '+__version__)
    subparsers = parser.add_subparsers(help='sub-command help')

    # create the parser for the "upload" command
    parser_upload = subparsers.add_parser('upload', help='Upload fastq files to ENA')
    parser_upload.add_argument('runs_file', help='Configuration file')
    parser_upload.add_argument('username', help='ENA username')
    parser_upload.add_argument('password', help='ENA password')
    parser_upload.set_defaults(func=upload_fastq_files)

    # create the parser for the "discover" command
    parser_discover = subparsers.add_parser('discover', help='Discover fastq files')
    parser_discover.add_argument('mode', choices=['single','paired'], help='Configuration file')
    parser_discover.add_argument('-1','--regex1', help='R1 regex pattern',required=True)
    parser_discover.add_argument('-2','--regex2', help='R2 regex pattern')
    parser_discover.set_defaults(func=disccover_files)

    # parse the args and call whatever function was selected
    args = parser.parse_args()
    if hasattr(args, 'func'):
        args.func(args)
    else:
        parser.print_help()


def upload_fastq_files(args):
    fastq_files = []
    for row in csv.DictReader(open(args.runs_file), delimiter="\t"):
        fastq_files.append(row["forward_file_name"])
        fastq_files.append(row["reverse_file_name"])
    
    url = "webin2.ebi.ac.uk"
    ftp = FTP(url, args.username, args.password)
    for fastq_file in tqdm(fastq_files):
        ftp.storbinary('STOR '+fastq_file, open(fastq_file, 'rb'))


def disccover_files(args):
    if args.mode=='SINGLE':
        samples = ff.find_single_fastq_samples(",",args.regex1)
    else:
        samples = ff.find_paired_fastq_samples(".",args.regex1, args.regex2)

    for sample in samples:
        sample.calculate_md5()

    write_template_files(samples)

def load_config():
    config_file = os.path.join(__package_dir__, "config.toml")
    config = toml.load(config_file)
    return config

def write_validator_sheet(ws: worksheet):
    config = load_config()
    ws.append(list(config['validators'].keys()))
    for i,category in enumerate(config['validators']):
        column = alphabet[i]
        for j,val in enumerate(config['validators'][category]):
            rownumber = j+2
            cell = ws[f"{column}{rownumber}"]
            cell.value = val


def get_validator_formulas():
    config = load_config()
    formulas = {}

    for i,key in enumerate(config['validators']):
        formulas[key] = f"{quote_sheetname('Validators')}!${alphabet[i]}$2:${alphabet[i]}${len(config['validators'][key])+1}"

    return formulas


def write_template_files(samples: List[Union[SingleSample,PairedSample]]):
    config = load_config()

    # Create a workbook and add a worksheet.
    wb = Workbook()
    
    # rename first sheet to samples
    ws1 = wb.active
    ws1.title = "Samples"
    ws2 = wb.create_sheet(title="Runs")
    ws3 = wb.create_sheet(title="Validators")

    write_validator_sheet(ws3)
    validator_formulas = get_validator_formulas()
    

    # Add a header at row 3
    ws1.append(["#checklist_accession","ERC000011"])
    ws1.append(["#unique_name_prefix"])

    # Add the units row
    ws1.append(["sample_alias","tax_id","scientific_name","sample_title","sample_description", "collection date", "geographic location (country and/or sea)"])
    ws1.append(["#units"])
    
    for sample in samples:
        ws1.append([sample.prefix, "", "", sample.prefix, sample.prefix])

    dv1 = DataValidation(type="list", formula1=validator_formulas['geographic location (country and/or sea)'], allow_blank=False, showErrorMessage=True)
    dv1.error ='Your entry is not in the list'
    dv1.errorTitle = 'Invalid Entry'
    ws1.add_data_validation(dv1)
    dv1.add(f"G5:G{len(samples)+4}")

    # must be 4 digits
    dv2 = DataValidation(type="list", formula1=validator_formulas['collection date'], allow_blank=False, showErrorMessage=True)
    dv2.error ='Your entry is not a year'
    dv2.errorTitle = 'Invalid Entry'
    ws1.add_data_validation(dv2)
    dv2.add(f"F5:F{len(samples)+4}")

    
    # Runs sheet
    wb.active = ws2
    ws2.append(["FileType","fastq","Read submission file type"])
    ws2.append(["sample","study","instrument_model","library_name","library_source","library_selection","library_strategy","library_layout","forward_file_name","forward_file_md5","reverse_file_name","reverse_file_md5"])

    for sample in samples:
        ws2.append([sample.prefix,'','',sample.prefix,'','','',os.path.split(sample.r1[0])[-1],sample.md5r1[0],os.path.split(sample.r2[0])[-1],sample.md5r2[0]])
    
    dv3 = DataValidation(type="list", formula1=validator_formulas['instrument_model'], allow_blank=False, showErrorMessage=True)
    dv3.error ='Your entry is not in the list'
    dv3.errorTitle = 'Invalid Entry'
    ws2.add_data_validation(dv3)
    dv3.add(f"C3:C{len(samples)+2}")

    dv4 = DataValidation(type="list", formula1=validator_formulas['library_source'], allow_blank=False, showErrorMessage=True)
    dv4.error ='Your entry is not in the list'
    dv4.errorTitle = 'Invalid Entry'
    ws2.add_data_validation(dv4)
    dv4.add(f"E3:E{len(samples)+2}")

    dv5 = DataValidation(type="list", formula1=validator_formulas['library_selection'], allow_blank=False, showErrorMessage=True)
    dv5.error ='Your entry is not in the list'
    dv5.errorTitle = 'Invalid Entry'
    ws2.add_data_validation(dv5)
    dv5.add(f"F3:F{len(samples)+2}")

    dv6 = DataValidation(type="list", formula1=validator_formulas['library_strategy'], allow_blank=False, showErrorMessage=True)
    dv6.error ='Your entry is not in the list'
    dv6.errorTitle = 'Invalid Entry'
    ws2.add_data_validation(dv6)
    dv6.add(f"G3:G{len(samples)+2}")

    


    # Save the file
    wb.save("samples.xlsx")


